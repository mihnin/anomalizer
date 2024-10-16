import pandas as pd
import numpy as np
from anomaly_detection import detect_anomalies, calculate_stats
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def process_file(file_path):
    # Чтение файла
    df = pd.read_excel(file_path)
    
    # Список для хранения результатов
    results = []

    # Проход по каждому столбцу
    for column in df.columns:
        if pd.api.types.is_numeric_dtype(df[column]):
            # Вычисление статистик
            iqr, q1, q3 = calculate_stats(df, column)
            lower_threshold = q1 - 1.5 * iqr
            upper_threshold = q3 + 1.5 * iqr

            # Нахождение аномалий
            anomalies = detect_anomalies(df, column, lower_threshold, upper_threshold)
            results.append({
                'column': column,
                'anomalies': anomalies,
                'lower_threshold': lower_threshold,
                'upper_threshold': upper_threshold
            })

    return results

def display_results(results):
    for result in results:
        column = result['column']
        anomalies = result['anomalies']
        print(f"Столбец: {column}")
        print(f"Аномалии:\n{anomalies}\n")

def create_anomalies_excel(results):
    # Создаем новый DataFrame для всех аномалий
    all_anomalies = pd.DataFrame()

    for result in results:
        column = result['column']
        anomalies = result['anomalies']
        lower_threshold = result['lower_threshold']
        upper_threshold = result['upper_threshold']
        
        # Добавляем столбец с флагом аномалии
        anomalies[f'{column}_is_anomaly'] = True
        
        # Объединяем с общим DataFrame
        if all_anomalies.empty:
            all_anomalies = anomalies
        else:
            all_anomalies = pd.merge(all_anomalies, anomalies, how='outer')

    # Создаем Excel-файл
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        all_anomalies.to_excel(writer, sheet_name='All Anomalies', index=False)
        
        # Получаем рабочий лист
        workbook = writer.book
        worksheet = workbook['All Anomalies']
        
        # Заполняем красным цветом ячейки с аномалиями
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        for col in worksheet.columns:
            if col[0].value.endswith('_is_anomaly'):
                for cell in col[1:]:
                    if cell.value:
                        worksheet.cell(row=cell.row, column=cell.column-1).fill = red_fill

    return output.getvalue()

# Пример использования
if __name__ == "__main__":
    file_path = 'path_to_your_file.xlsx'
    results = process_file(file_path)
    display_results(results)
